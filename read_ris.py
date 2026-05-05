import openpyxl

wb = openpyxl.load_workbook('ris.xlsx')
ws = wb.active

print('Sheet name:', ws.title)
print('Max row:', ws.max_row)
print('Max col:', ws.max_column)
print('\nAll rows with content:')

for i in range(1, ws.max_row + 1):
    row_data = []
    for j in range(1, ws.max_column + 1):
        cell = ws.cell(row=i, column=j)
        row_data.append(f"{j}:{cell.value}")
    print(f"Row {i}: {row_data}")
